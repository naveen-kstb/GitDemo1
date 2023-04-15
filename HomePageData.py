import openpyxl


class HomePageData:
    test_HomePage_data = [{"firstname": "Naveen", "email": "hello@gmail.com", "gender": "Male"},
                          {"firstname": "Saisha", "email": "hello@gmail.com", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook(r"C:\\Users\\naveen.kstb\\Desktop\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # To get Rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # To get Columns
                    # print(sheet.cell(row=i, column=j).value)

                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
