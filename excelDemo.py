import openpyxl

book = openpyxl.load_workbook(r"C:\\Users\\naveen.kstb\\Desktop\\PythonDemo.xlsx")
sheet = book.active  # sheet: <Worksheet "Sheet1">
Dict = {}  # Dict: <class 'dict'>: {'firstname': 'Saisha', 'Lastname': 'Culli'}
cell = sheet.cell(row=2, column=2)  # cell: <cell 'Sheet1'.B1>
print(cell.value)

sheet.cell(row=2, column=3).value = "Hello@gmail.com"
print(sheet.cell(row=2, column=3).value)

print(sheet.max_row)

print(sheet.max_column)

print(sheet["A3"].value)

for i in range(1, sheet.max_row + 1):  # To get Rows
    if sheet.cell(row=i, column=1).value == "Testcase 1":

        for j in range(2, sheet.max_column + 1):  # To get Columns
            print(sheet.cell(row=i, column=j).value)

            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict)
